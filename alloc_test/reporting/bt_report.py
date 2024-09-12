import os
from dash.dependencies import Input, Output
import seaborn as sns
import shap
import numpy as np
import pandas as pd
from waitress import serve
import webbrowser
from dash import dcc, html
from dash import dash_table
import dash
import plotly.graph_objs as go
import pyfolio as pf
from matplotlib import pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

class PyfolioReport:
    def __init__(self,output_dir):
        self.output_dir=output_dir

    def _generate_pyfolio_report(self, strategy_name, portfolio_returns, positions, transactions):
        strategy_output_dir = os.path.join(self.output_dir, strategy_name)
        os.makedirs(strategy_output_dir, exist_ok=True)

        plt.figure(figsize=(15, 10), dpi=100)

        with PdfPages(f'{strategy_output_dir}/pyfolio_report.pdf') as pdf:
            pf.create_full_tear_sheet(portfolio_returns)
            pdf.savefig()
            plt.close()

        perf_stats = pf.timeseries.perf_stats(portfolio_returns)
        perf_stats.to_csv(f'{strategy_output_dir}/performance_metrics.csv')

        plt.figure(figsize=(15, 10), dpi=100)
        pf.create_returns_tear_sheet(portfolio_returns)
        plt.savefig(f'{strategy_output_dir}/returns_plot.png')
        plt.close()

        plt.figure(figsize=(15, 10), dpi=100)
        pf.create_position_tear_sheet(portfolio_returns, positions)
        plt.savefig(f'{strategy_output_dir}/positions_plot.png')
        plt.close()

        plt.figure(figsize=(15, 10), dpi=100)
        pf.create_txn_tear_sheet(portfolio_returns, positions, transactions)
        plt.savefig(f'{strategy_output_dir}/transactions_plot.png')
        plt.close()

        plt.figure(figsize=(15, 10), dpi=100)
        pf.create_interesting_times_tear_sheet(portfolio_returns)
        plt.savefig(f'{strategy_output_dir}/interesting_times_plot.png')
        plt.close()

    def _generate_heatmap(self, asset_returns):
        # Add asset correlation matrix heatmap based on returns
        plt.figure(figsize=(10, 8))
        sns.heatmap(asset_returns.corr(), annot=True, cmap="coolwarm", fmt=".2f")
        plt.title('Asset Returns Correlation Heatmap')
        plt.savefig(f'{self.output_dir}/correlation_heatmap.png')
        plt.close()
        print(f"Pyfolio report and plots saved to '{self.output_dir}' directory.")


class ShapReport:
    def __init__(self, output_dir):
        self.output_dir=output_dir

    def generate_shap_explainability(self, strategy_name, asset_returns):
        strategy_output_dir = os.path.join(self.output_dir, strategy_name)
        os.makedirs(strategy_output_dir, exist_ok=True)

        plt.figure(figsize=(15, 10), dpi=100)
        explainer = shap.Explainer(strategy_name)
        shap_values = explainer(asset_returns)
        shap.summary_plot(shap_values, asset_returns)


class DashReport:
    def __init__(self, asset_returns, strategy_results, benchmark_returns, port):
        self.asset_returns=asset_returns
        self.strategy_results = strategy_results
        self.benchmark_returns = benchmark_returns
        self.strategy_results_dash = self._generate_performance_report()
        self.app = dash.Dash(__name__)
        self._setup_layout()
        self._setup_callbacks()
        self.port = port


    def _generate_performance_report(self):
        strategy_results_dash = {}
        for name, results in self.strategy_results.items():
            self.rolling_metrics = {
                'rolling_sharpe': pf.timeseries.rolling_sharpe(results['portfolio_returns'], 126),
                'rolling_beta': pf.timeseries.rolling_beta(results['portfolio_returns'], self.benchmark_returns, 126),
                'portfolio_cumulative_returns': (1 + results['portfolio_returns']).cumprod() - 1,
                'weights': results['weights'], 'asset_prices':results['asset_prices'], 'positions': results['positions'],
                'portfolio_pnl': results['portfolio_pnl'], 'asset_pnl': results['asset_pnl'],
                'portfolio_cum_pnl':results['portfolio_pnl'].cumsum(),'asset_cum_pnl': results['asset_pnl'].cumsum(),
                'asset_cumulative_returns': (1 + results['asset_returns']).cumprod() - 1
            }
            self.performance_metrics = {
                'best_opti_algo' : results['best_opti_algo'],
                'best_params':str(results['best_params']),
                'annual_return': pf.timeseries.annual_return(results['portfolio_returns']),
                'annual_volatility': pf.timeseries.annual_volatility(results['portfolio_returns']),
                'sharpe_ratio': pf.timeseries.sharpe_ratio(results['portfolio_returns']),
                'calmar_ratio': pf.timeseries.calmar_ratio(results['portfolio_returns']),
                'max_drawdown': pf.timeseries.max_drawdown(results['portfolio_returns']),
                'omega_ratio': pf.timeseries.omega_ratio(results['portfolio_returns']),
                'sortino_ratio': pf.timeseries.sortino_ratio(results['portfolio_returns']),
                'tail_ratio': pf.timeseries.tail_ratio(results['portfolio_returns']),
                'daily_var': pf.timeseries.value_at_risk(results['portfolio_returns'])
            }
            self.performance_metrics = {k: (0 if pd.isna(v) else v) for k, v in self.performance_metrics.items()}
            strategy_results_dash[name] = {**self.rolling_metrics, **self.performance_metrics}
        return strategy_results_dash

    def _prepare_data_for_table(self):
        return [{'Strategy': s, **{k: round(v, 2) if isinstance(v, (float, int)) else v
                                   for k, v in m.items() if k in list(self.performance_metrics.keys())}}
                for s, m in self.strategy_results_dash.items()]

    def _generate_correlation_heatmap(self):
        # Generate the correlation matrix and create a static heatmap
        correlation_matrix =self.asset_returns.corr()
        heatmap = {
            'data': [
                go.Heatmap(
                    z=correlation_matrix.values,
                    x=correlation_matrix.columns,
                    y=correlation_matrix.index,
                    colorscale='Viridis',
                    colorbar={'title': 'Correlation'},
                    text=correlation_matrix.values,  # Set the text to show correlation values
                    texttemplate="%{text:.2f}",  # Format the numbers to 2 decimal places
                    hoverinfo='z',  # Only show the z value (correlation) on hover
                )
            ],
            'layout': go.Layout(
                title='Asset Returns Correlation Heatmap',
                xaxis={'title': 'Assets'},
                yaxis={'title': 'Assets'}
            )
        }
        return heatmap

    def _setup_layout(self):
        self.app.layout = html.Div([
            html.H1('Strategies Static Performance Metrics'),
            dash_table.DataTable(id='performance-metrics-table',
                columns=[{"name": i, "id": i} for i in self._prepare_data_for_table()[0].keys()],
                data=self._prepare_data_for_table(), style_table={'overflowX': 'auto'},
                style_header={'backgroundColor': 'rgb(230, 230, 230)', 'fontWeight': 'bold'},
                style_cell={'textAlign': 'center'}),

            html.H1('Strategies Rolling Performance Metrics'),
            *[dcc.Graph(id=f'graph-{metric}', figure={'data': [go.Scatter(x=metrics.index, y=metrics, mode='lines', name=strategy)
                for strategy, data in self.strategy_results_dash.items() for metrics in [data[metric]]], 'layout': go.Layout(title=f'{metric.capitalize()} Comparison Across Strategies', xaxis={'title': 'Date'}, yaxis={'title': metric.replace("_", " ").title()})})
              for metric in ['rolling_sharpe', 'rolling_beta', 'portfolio_cumulative_returns', 'portfolio_pnl', 'portfolio_cum_pnl']],

            html.H1('Asset-Level Metrics Per Strategy'),
            dcc.Dropdown(id='strategy-dropdown', options=[{'label': s, 'value': s} for s in self.strategy_results_dash.keys()],
                         value=list(self.strategy_results_dash.keys())[0], clearable=False, style={'width': '50%'}),
            *[dcc.Graph(id=f'asset-{metric}-graph') for metric in
              ['weights', 'asset_prices', 'positions', 'asset_pnl','asset_cum_pnl', 'asset_cumulative_returns']],
            # Add asset correlation heatmap graph
            html.H1('Asset Correlation Heatmap'),
            dcc.Graph(id='correlation-heatmap', figure=self._generate_correlation_heatmap())  # Use the static heatmap figure
        ])

    def _setup_callbacks(self):
        @self.app.callback(
            [Output(f'asset-{metric}-graph', 'figure') for metric in ['weights',
                                                'asset_prices', 'positions', 'asset_pnl', 'asset_cum_pnl', 'asset_cumulative_returns']],
            [Input('strategy-dropdown', 'value')]
        )
        def update_asset_graphs(selected_strategy):
            figures = [
                {
                    'data': [
                        go.Scatter(x=metric.index, y=metric[symbol], mode='lines', name=symbol)
                        for symbol in metric.columns
                    ],
                    'layout': go.Layout(
                        title=f'{metric_name.capitalize()} for {selected_strategy}',
                        xaxis={'title': 'Date'},
                        yaxis={'title': metric_name.replace("_", " ").title()}
                    )
                }
                for metric_name, metric in self.strategy_results_dash[selected_strategy].items()
                if metric_name in ['weights','asset_prices', 'positions', 'asset_pnl','asset_cum_pnl', 'asset_cumulative_returns']
            ]

            return figures
    def run_server(self):
        webbrowser.open(f"http://127.0.0.1:{self.port}")
        serve(self.app.server, host='0.0.0.0', port=self.port)

class ExcelReport:

    def calculate_maximum_drawdown(self, report):
        cum_pnl = report['pnl'].cumsum()
        dd = cum_pnl.cummax() - cum_pnl
        max_dd = dd.max() / cum_pnl.max()
        max_dd_date = dd.idxmax()

        # Further drawdown analysis
        max_dd_period = (dd != 0).astype(int).groupby(dd.eq(0).cumsum()).cumsum().max()
        recov_date = dd[dd == 0][max_dd_date:].index.min() if not dd[dd == 0][max_dd_date:].empty else dd.index[-1]
        recov_duration = (recov_date - max_dd_date).days
        max_dd_duration = max_dd_period

        return {
            'max_drawdown': max_dd,
            'max_drawdown_date': max_dd_date,
            'max_drawdown_duration': max_dd_duration,
            'recovery_date': recov_date,
            'recovery_duration': recov_duration
        }

    def compute_rolling_metrics(self, report):
        rolling_window = 252  # 1 year rolling window
        rolling_sharpe = np.sqrt(252) * report['pnl'].rolling(window=rolling_window).mean() / report['pnl'].rolling(window=rolling_window).std()
        rolling_sortino = np.sqrt(252) * report['pnl'].rolling(window=rolling_window).mean() / report['pnl'][report['pnl'] < 0].rolling(window=rolling_window).std()
        hit_ratio = (report['pnl'] > 0).rolling(window=rolling_window).mean()
        return rolling_sharpe, rolling_sortino, hit_ratio

    def generate_excel_report(self, df_list, table_names_list, strat_names, output_path):
        # Initialize Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        invisible_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF'))

        # Create sheets for each strategy and summary
        for name in strat_names:
            wb.create_sheet(title=name)

        # Populate the "Summary" sheet
        summary_ws = wb["Summary"]
        summary_ws.append(["Strategy", "Sharpe", "HitRatio", "Sortino", "Max Drawdown", "Recov Duration", "Start Date DD", "End Date DD", "Reco Date DD"])

        for idx, df in enumerate(df_list):
            strat_name = strat_names[idx]
            rolling_sharpe, rolling_sortino, hit_ratio = self.compute_rolling_metrics(df)

            # Calculate metrics for summary
            sharpe = rolling_sharpe.mean()
            sortino = rolling_sortino.mean()
            hit = hit_ratio.mean()
            max_dd_info = self.calculate_maximum_drawdown(df)

            summary_ws.append([strat_name, sharpe, hit, sortino, max_dd_info['max_drawdown'], max_dd_info['recovery_duration'], max_dd_info['max_drawdown_date'], max_dd_info['recovery_date']])

            # Populate each strategy sheet
            strategy_ws = wb[strat_name]
            self.write_dataframes_to_excel([df], strategy_ws, table_names_list, strat_name)

        # Save the workbook
        wb.save(output_path)
        print(f"Excel report saved to {output_path}")

    def write_dataframes_to_excel(self, dataframes, ws, table_names_list, strat):
        invisible_border = Border(
            left=Side(style='thin', color=None),
            right=Side(style='thin', color=None),
            top=Side(style='thin', color=None),
            bottom=Side(style='thin', color=None))

        used_cell_border_medium = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))

        used_cell_border_thin = Border(left=Side(style='thin'),
                                       right=Side(style='thin'),
                                       top=Side(style='thin'),
                                       bottom=Side(style='thin'))

        def apply_color(cell, value):
            if value < 0:
                cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')  # Light Red
            elif value >= 0:
                cell.fill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')  # Light Blue

        start_row = 1
        start_col = 1
        for index, df in enumerate(dataframes):
            title_cell = ws.cell(row=start_row, column=start_col, value=str(table_names_list[index]))
            title_cell.font = Font(bold=True)
            title_cell.alignment = Alignment(horizontal='left')
            start_row += 2

            for j, header in enumerate(df.columns, start=start_col + 1):
                cell = ws.cell(row=start_row, column=j, value=header)
                cell.border = used_cell_border_medium

            for r_idx, (idx, row) in enumerate(df.iterrows(), start=start_row + 1):
                cell = ws.cell(row=r_idx, column=start_col, value=idx)
                cell.border = used_cell_border_medium
                for c_idx, value in enumerate(row.values, start=start_col + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = np.round(value, 2) if isinstance(value, float) else value
                    apply_color(cell, value)
                    cell.border = used_cell_border_thin
            start_row += len(df) + 3  # Add space for the next DataFrame

    def plot_performance(self, strategy_results, output_dir):
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        for strategy, metrics in strategy_results.items():
            fig, axs = plt.subplots(2, 2, figsize=(12, 8))

            # Cumulative PnL plot
            axs[0][0].plot(metrics['report'].set_index('date')['pnl'].cumsum(), label='Cumulative PnL')
            axs[0][0].legend(loc='lower right')
            axs[0][0].set_title('Cumulative PnL')

            # Rolling Sharpe ratio plot
            axs[0][1].plot(metrics['rolling_sharpe'], label='Rolling Sharpe Ratio')
            axs[0][1].set_title('Rolling 1Y Sharpe Ratio')

            # Rolling Sortino ratio plot
            axs[1][0].plot(metrics['rolling_sortino'], label='Rolling Sortino Ratio')
            axs[1][0].set_title('Rolling 1Y Sortino Ratio')

            # Hit Ratio plot
            axs[1][1].plot(metrics['hit_ratio'], label='Rolling Hit Ratio')
            axs[1][1].set_title('Rolling 1Y Hit Ratio')

            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, f"{strategy}_performance.png"))
            plt.close()
